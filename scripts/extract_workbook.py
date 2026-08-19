from pathlib import Path
import json
import re
from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"#NAME?", "None"}:
        return ""
    return text


def rows_for_sheet(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = [clean(v) for v in row]
        while values and values[-1] == "":
            values.pop()
        if any(values):
            rows.append(values)
    return rows


def parse_caravan(rows):
    intro = [r[0] for r in rows[:9] if r and r[0]]
    levels = []
    for row in rows:
        if len(row) >= 4 and str(row[1]).isdigit():
            levels.append({"level": row[1], "start": row[2], "end": row[3] if len(row) > 3 else ""})
    return {"intro": intro, "levels": levels}


def parse_duel(rows):
    entries = []
    for row in rows:
        if len(row) >= 4 and (row[1].startswith("Day") or "공지" in row[1]):
            entries.append({"date": row[1], "ko": row[2], "en": row[3]})
    return entries


def parse_member_rows(rows):
    member_rows = []
    for row in rows:
        if len(row) >= 11 and row[3]:
            member_rows.append({
                "status": row[0],
                "current": row[1],
                "next": row[2],
                "name": row[3],
                "day1": row[4],
                "day2": row[5],
                "day3": row[6],
                "day4": row[7],
                "day5": row[8],
                "day6": row[9],
                "total": row[10],
                "note": row[11] if len(row) > 11 else ""
            })
    return member_rows


def main():
    base = Path.home() / "Downloads"
    workbook = next(base.glob("*LIr*/*.xlsx"))
    wb = load_workbook(workbook, data_only=True)
    data = {"source": workbook.name, "sheets": {}}

    for ws in wb.worksheets:
        rows = rows_for_sheet(ws)
        if ws.title == "연맹 대결 가이드":
            data["sheets"][ws.title] = {"kind": "duel", "entries": parse_duel(rows)}
        elif ws.title == "캐러밴 표":
            data["sheets"][ws.title] = {"kind": "caravan", **parse_caravan(rows)}
        elif ws.title == "시트14":
            data["sheets"][ws.title] = {"kind": "members", "rows": parse_member_rows(rows)}
        else:
            data["sheets"][ws.title] = {"kind": "raw", "rows": rows}

    out = Path(__file__).resolve().parents[1] / "src" / "workbook-data.json"
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
